"""
Planilha de Ligações -- pedido do Clayton (2026-09-03): uma tela pra
controlar as ligações de prospecção (dia, empresa, com quem falou, pra
quem terceirizam, quem é o responsável pela área de suplementos/novos
produtos/contratação de fabricantes), editável direto no sistema, com
histórico completo e exportação em Excel/PDF.

Compartilhada pela empresa toda (mesmo padrão de respostas prontas e
etiquetas) -- qualquer um com acesso às conversas pode ver e editar.

Lembrete de "entrar em contato novamente" (2026-09-03): cada linha pode
ter uma data de próximo contato; quando chega o dia, o Assistente Seja
Alpha avisa quem criou a linha no chat interno, com opção de prorrogar
o aviso (ver avisar_ligacoes_pendentes_se_preciso, chamada pelo
agendador em scheduler.py -- mesmo padrão do aviso de conversa parada).
"""
import datetime
import io

from flask import Blueprint, g, jsonify, request, send_file

from ..context import ApiError, get_db, requires_auth

bp = Blueprint("ligacoes", __name__, url_prefix="/api/v1/ligacoes")

CAMPOS_EDITAVEIS = (
    "data_ligacao", "empresa_contatada", "contato_nome", "telefone",
    "email", "data_envio_email", "terceiriza_para", "responsavel_area",
    "observacoes", "proximo_contato_em", "aceitacao", "negociacao_fechada",
)

ACEITACAO_VALIDAS = ("quente", "morno", "frio")

COLUNAS = (
    ("data_ligacao", "Data"),
    ("empresa_contatada", "Empresa"),
    ("contato_nome", "Com quem falei"),
    ("telefone", "Telefone"),
    ("email", "E-mail"),
    ("data_envio_email", "Data envio e-mail"),
    ("terceiriza_para", "Terceirizam com"),
    ("responsavel_area", "Responsável (suplementos/novos produtos/fabricantes)"),
    ("proximo_contato_em", "Próximo contato"),
    ("aceitacao", "Aceitação"),
    ("negociacao_fechada", "Negociação fechada"),
    ("observacoes", "Observações"),
)


def _now_iso():
    return datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"


def _carregar(conn, empresa_id, ligacao_id):
    row = conn.execute(
        "SELECT * FROM crm_ligacoes WHERE id = ? AND empresa_id = ?", (ligacao_id, empresa_id)
    ).fetchone()
    if row is None:
        raise ApiError("Ligação não encontrada.", status=404, codigo="nao_encontrado")
    return row


@bp.get("")
@requires_auth
def listar():
    conn = get_db()
    rows = conn.execute(
        "SELECT l.*, u.nome AS criado_por_nome FROM crm_ligacoes l "
        "LEFT JOIN usuarios u ON u.id = l.criado_por "
        "WHERE l.empresa_id = ? ORDER BY l.ordem, l.id",
        (g.empresa_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.post("")
@requires_auth
def criar():
    conn = get_db()
    usuario = g.usuario_atual
    agora = _now_iso()
    maior_ordem = conn.execute(
        "SELECT COALESCE(MAX(ordem), -1) AS v FROM crm_ligacoes WHERE empresa_id = ?", (g.empresa_id,)
    ).fetchone()["v"]
    cur = conn.execute(
        "INSERT INTO crm_ligacoes (empresa_id, data_ligacao, ordem, criado_por, criado_em) VALUES (?, ?, ?, ?, ?)",
        (g.empresa_id, datetime.date.today().isoformat(), maior_ordem + 1, usuario["id"], agora),
    )
    conn.commit()
    return jsonify(dict(_carregar(conn, g.empresa_id, cur.lastrowid))), 201


@bp.put("/<int:ligacao_id>")
@requires_auth
def atualizar(ligacao_id):
    conn = get_db()
    _carregar(conn, g.empresa_id, ligacao_id)  # 404 se não for desta empresa
    dados = request.get_json(silent=True) or {}
    campos, valores = [], []
    for campo in CAMPOS_EDITAVEIS:
        if campo in dados:
            valor = dados[campo]
            if campo == "aceitacao" and valor not in ACEITACAO_VALIDAS and valor not in (None, ""):
                raise ApiError("Aceitação inválida — use quente, morno, frio ou deixe em branco.", status=400)
            campos.append(f"{campo} = ?")
            valores.append((valor or "").strip() or None if isinstance(valor, str) else valor)
    # Mudou a data do próximo contato: libera pra avisar de novo (senão
    # marcar uma data nova nunca dispararia, porque aviso_enviado_em
    # ficaria de uma rodada anterior).
    if "proximo_contato_em" in dados:
        campos.append("aviso_enviado_em = NULL")
    if not campos:
        return jsonify({"ok": True})
    campos.append("atualizado_em = ?")
    valores.append(_now_iso())
    campos.append("atualizado_por = ?")
    valores.append(g.usuario_atual["id"])
    valores.extend([ligacao_id, g.empresa_id])
    conn.execute(f"UPDATE crm_ligacoes SET {', '.join(campos)} WHERE id = ? AND empresa_id = ?", valores)
    conn.commit()
    return jsonify(dict(_carregar(conn, g.empresa_id, ligacao_id)))


@bp.post("/<int:ligacao_id>/prorrogar")
@requires_auth
def prorrogar(ligacao_id):
    """Adia o próximo contato pelos dias configurados em
    dias_prorrogar_ligacao (Configuração), contados de HOJE -- pra
    quando a pessoa vê o aviso e quer empurrar de novo."""
    from .. import whatsapp_service

    conn = get_db()
    _carregar(conn, g.empresa_id, ligacao_id)
    config = whatsapp_service.obter_configuracao(conn, g.empresa_id)
    dias = config.get("dias_prorrogar_ligacao") or 3
    nova_data = (datetime.date.today() + datetime.timedelta(days=dias)).isoformat()
    conn.execute(
        "UPDATE crm_ligacoes SET proximo_contato_em = ?, aviso_enviado_em = NULL, "
        "vezes_prorrogado = vezes_prorrogado + 1, atualizado_em = ?, atualizado_por = ? WHERE id = ?",
        (nova_data, _now_iso(), g.usuario_atual["id"], ligacao_id),
    )
    conn.commit()
    return jsonify(dict(_carregar(conn, g.empresa_id, ligacao_id)))


@bp.delete("/<int:ligacao_id>")
@requires_auth
def excluir(ligacao_id):
    conn = get_db()
    _carregar(conn, g.empresa_id, ligacao_id)
    conn.execute("DELETE FROM crm_ligacoes WHERE id = ? AND empresa_id = ?", (ligacao_id, g.empresa_id))
    conn.commit()
    return jsonify({"ok": True})


def avisar_ligacoes_pendentes_se_preciso(conn):
    """Chamado periodicamente pelo agendador (ver scheduler.py). Avisa
    quem criou a linha, no chat interno, quando chega (ou já passou) o
    dia marcado pra entrar em contato de novo -- uma vez por dia,
    enquanto ninguém prorrogar ou trocar a data."""
    from .. import chat_interno_service, followup_service

    empresas = conn.execute(
        "SELECT empresa_id FROM configuracoes_whatsapp WHERE aviso_ligacoes_ativo = 1"
    ).fetchall()
    if not empresas:
        return 0
    hoje = datetime.date.today().isoformat()
    total = 0
    for emp in empresas:
        empresa_id = emp["empresa_id"]
        remetente_id = followup_service._remetente_do_sistema(conn, empresa_id)
        if remetente_id is None:
            continue
        pendentes = conn.execute(
            "SELECT id, empresa_contatada, contato_nome, criado_por, proximo_contato_em "
            "FROM crm_ligacoes WHERE empresa_id = ? AND proximo_contato_em IS NOT NULL "
            "AND proximo_contato_em <= ? "
            "AND (aviso_enviado_em IS NULL OR substr(aviso_enviado_em, 1, 10) != ?)",
            (empresa_id, hoje, hoje),
        ).fetchall()
        for lig in pendentes:
            if not lig["criado_por"] or lig["criado_por"] == remetente_id:
                continue
            destino = conn.execute(
                "SELECT id, setor FROM usuarios WHERE id = ? AND ativo = 1", (lig["criado_por"],)
            ).fetchone()
            if destino is None:
                continue
            nome_empresa = lig["empresa_contatada"] or lig["contato_nome"] or "um cliente"
            texto = (
                "🔔 Lembrete: hoje é o dia marcado pra entrar em contato de novo com *" + str(nome_empresa) + "*"
                + (f" ({lig['contato_nome']})" if lig["contato_nome"] and lig["empresa_contatada"] else "")
                + ". Não esqueça de ligar! Se quiser adiar, abra Ligações e clique em \"Prorrogar\" nessa linha."
            )
            conversa_interna_id = chat_interno_service.buscar_conversa_existente(conn, remetente_id, destino["id"])
            if conversa_interna_id:
                chat_interno_service.reabrir_conversa(conn, conversa_interna_id)
                chat_interno_service.enviar_mensagem(conn, conversa_interna_id, remetente_id, texto)
            else:
                chat_interno_service.iniciar_conversa(conn, remetente_id, destino["id"], destino["setor"], texto)
            conn.execute("UPDATE crm_ligacoes SET aviso_enviado_em = ? WHERE id = ?", (_now_iso(), lig["id"]))
            total += 1
    return total


def _linhas_ordenadas(conn):
    return conn.execute(
        "SELECT * FROM crm_ligacoes WHERE empresa_id = ? ORDER BY ordem, id", (g.empresa_id,)
    ).fetchall()


@bp.get("/exportar.xlsx")
@requires_auth
def exportar_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    conn = get_db()
    linhas = _linhas_ordenadas(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ligações"
    cabecalho = [rotulo for _, rotulo in COLUNAS]
    ws.append(cabecalho)
    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="0A7D67")
    rotulos_aceitacao = {"quente": "🔥 Quente", "morno": "🟡 Morno", "frio": "❄️ Frio"}
    for linha in linhas:
        valores_linha = []
        for campo, _ in COLUNAS:
            if campo == "aceitacao":
                valores_linha.append(rotulos_aceitacao.get(linha[campo], ""))
            elif campo == "negociacao_fechada":
                valores_linha.append("Sim" if linha[campo] else "")
            else:
                valores_linha.append(linha[campo] or "")
        ws.append(valores_linha)
    larguras = [12, 26, 20, 16, 22, 16, 22, 34, 14, 12, 16, 34]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True, download_name="ligacoes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/exportar.pdf")
@requires_auth
def exportar_pdf():
    from fpdf import FPDF

    conn = get_db()
    linhas = _linhas_ordenadas(conn)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Ligações", ln=1)
    pdf.set_font("Helvetica", "", 9)

    # Larguras simples com texto cortado (com "…" se passar do espaço) --
    # tentar quebrar linha em várias alturas por célula com fpdf2 dá bug
    # de alinhamento fácil; pra uma exportação de apoio, previsível e
    # sem quebrar é melhor que bonito.
    larguras = [18, 24, 22, 20, 24, 18, 24, 34, 16, 16, 20, 32]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(10, 125, 103)
    pdf.set_text_color(255, 255, 255)
    for (campo, rotulo), largura in zip(COLUNAS, larguras):
        pdf.cell(largura, 8, rotulo, border=1, fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)

    def _cortar(texto, largura_mm):
        max_chars = max(6, int(largura_mm / 1.7))
        texto = str(texto or "")
        # A fonte padrão (Helvetica) só cobre Latin-1 -- troca qualquer
        # caractere fora disso (emoji etc) por "?" em vez de quebrar a
        # exportação inteira por causa de um caractere solto.
        texto = texto.encode("latin-1", "replace").decode("latin-1")
        return texto if len(texto) <= max_chars else texto[: max_chars - 1] + "..."

    rotulos_aceitacao_pdf = {"quente": "Quente", "morno": "Morno", "frio": "Frio"}
    for linha in linhas:
        for (campo, _), largura in zip(COLUNAS, larguras):
            if campo == "aceitacao":
                texto_celula = rotulos_aceitacao_pdf.get(linha[campo], "")
            elif campo == "negociacao_fechada":
                texto_celula = "Sim" if linha[campo] else ""
            else:
                texto_celula = linha[campo]
            pdf.cell(largura, 7, _cortar(texto_celula, largura), border=1)
        pdf.ln()

    saida = bytes(pdf.output())
    buffer = io.BytesIO(saida)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="ligacoes.pdf", mimetype="application/pdf")
